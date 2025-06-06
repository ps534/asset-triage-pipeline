"""
Spreadsheet output generation system.
Creates filtered worksheet with proper formatting and columns.
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
import shutil
import logging
from datetime import datetime

from src.pipeline.pipeline_orchestrator import ProcessingResult


class SpreadsheetGeneratorError(Exception):
    """Custom exception for spreadsheet generation errors."""
    pass


class SpreadsheetGenerator:
    """Generates the filtered output spreadsheet."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.output_columns = [
            'Asset Name', 'Company', 'Development Timeline', 'Indication', 'Pursue', 'Fail Reasons',
            'Degree of Unmet Need', 'Indication Match Found', 'Unmet Need Z-Score',
            'Repurposing (if from Prompt A)', 'Rationale', 'Error',
            'Novelty & Differentiation Score', 'Unmet Medical Need Score', 'Development Stage Score',
            'Capital Efficiency Score', 'Peak Sales Potential Score', 'IP Strength & Duration Score',
            'Probability Technical Success Score', 'Competitive Landscape Score', 'Transactability Score',
            'Regulatory Path Complexity Score', 'Strategic Fit Score',
            'Novelty & Differentiation Rationale', 'Unmet Medical Need Rationale', 'Development Stage Rationale',
            'Capital Efficiency Rationale', 'Peak Sales Potential Rationale', 'IP Strength & Duration Rationale',
            'Probability Technical Success Rationale', 'Competitive Landscape Rationale', 'Transactability Rationale',
            'Regulatory Path Complexity Rationale', 'Strategic Fit Rationale'
        ]
        
        # Define colors for conditional formatting
        self.colors = {
            'pursue': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light green
            'dont_pursue': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Light red
            'error': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Light yellow
            'header': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray
        }
    
    def create_backup(self, file_path: Path) -> Path:
        """Create backup copy of original file."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = file_path.parent / f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
            shutil.copy2(file_path, backup_path)
            self.logger.info(f"Created backup: {backup_path}")
            return backup_path
        except Exception as e:
            raise SpreadsheetGeneratorError(f"Failed to create backup: {str(e)}")
    
    def generate_filtered_sheet(self, input_file: Path, results: List[ProcessingResult]) -> None:
        """
        Generate new 'Filtered' worksheet in the existing workbook.
        Applies conditional formatting for pursue/don't pursue/error.
        """
        try:
            # Create backup first
            self.create_backup(input_file)
            
            # Convert results to DataFrame
            df = self._results_to_dataframe(results)
            
            if input_file.suffix.lower() == '.csv':
                # For CSV files, create a new Excel file
                output_path = input_file.parent / f"{input_file.stem}_filtered.xlsx"
                self._create_excel_from_csv(input_file, df, output_path)
            else:
                # For Excel files, add new sheet
                self._add_sheet_to_excel(input_file, df)
            
            self.logger.info(f"Generated filtered output with {len(df)} asset-indication pairs")
            
        except Exception as e:
            raise SpreadsheetGeneratorError(f"Failed to generate spreadsheet: {str(e)}")
    
    def _results_to_dataframe(self, results: List[ProcessingResult]) -> pd.DataFrame:
        """Convert ProcessingResult objects to DataFrame."""
        data = []
        
        self.logger.debug(f"Converting {len(results)} results to DataFrame")
        for i, result in enumerate(results):
            # Format fail reasons as comma-separated string
            fail_reasons_str = ", ".join(result.fail_reasons) if result.fail_reasons else ""
            
            # Format pursue field
            if result.pursue is None:
                pursue_str = "Error"
            elif result.pursue:
                pursue_str = "Yes"
            else:
                pursue_str = "No"
            
            # Format repurposing field
            repurposing_str = "Yes" if result.is_repurposing else "No"
            
            # Format match found field
            match_found_str = "Yes" if result.indication_match_found else "No"
            
            # Format z-score field
            z_score_str = f"{result.unmet_need_z_score:.2f}" if result.unmet_need_z_score is not None else ""
            
            row = {
                'Asset Name': result.asset_name,
                'Company': result.company_name,
                'Development Timeline': result.timeline or "Timeline research unavailable",
                'Indication': result.indication,
                'Pursue': pursue_str,
                'Fail Reasons': fail_reasons_str,
                'Degree of Unmet Need': result.degree_of_unmet_need,
                'Indication Match Found': match_found_str,
                'Unmet Need Z-Score': z_score_str,
                'Repurposing (if from Prompt A)': repurposing_str,
                'Rationale': result.rationale or "",
                'Error': result.error or "",
                'Novelty & Differentiation Score': result.novelty_differentiation_score if result.novelty_differentiation_score is not None else "N/A",
                'Unmet Medical Need Score': result.unmet_medical_need_score if result.unmet_medical_need_score is not None else "N/A",
                'Development Stage Score': result.development_stage_score if result.development_stage_score is not None else "N/A",
                'Capital Efficiency Score': result.capital_efficiency_score if result.capital_efficiency_score is not None else "N/A",
                'Peak Sales Potential Score': result.peak_sales_potential_score if result.peak_sales_potential_score is not None else "N/A",
                'IP Strength & Duration Score': result.ip_strength_duration_score if result.ip_strength_duration_score is not None else "N/A",
                'Probability Technical Success Score': result.probability_technical_success_score if result.probability_technical_success_score is not None else "N/A",
                'Competitive Landscape Score': result.competitive_landscape_score if result.competitive_landscape_score is not None else "N/A",
                'Transactability Score': result.transactability_score if result.transactability_score is not None else "N/A",
                'Regulatory Path Complexity Score': result.regulatory_path_complexity_score if result.regulatory_path_complexity_score is not None else "N/A",
                'Strategic Fit Score': result.strategic_fit_score if result.strategic_fit_score is not None else "N/A",
                'Novelty & Differentiation Rationale': result.novelty_differentiation_rationale if result.novelty_differentiation_rationale is not None else "Insufficient data",
                'Unmet Medical Need Rationale': result.unmet_medical_need_rationale if result.unmet_medical_need_rationale is not None else "Insufficient data",
                'Development Stage Rationale': result.development_stage_rationale if result.development_stage_rationale is not None else "Insufficient data",
                'Capital Efficiency Rationale': result.capital_efficiency_rationale if result.capital_efficiency_rationale is not None else "Insufficient data",
                'Peak Sales Potential Rationale': result.peak_sales_potential_rationale if result.peak_sales_potential_rationale is not None else "Insufficient data",
                'IP Strength & Duration Rationale': result.ip_strength_duration_rationale if result.ip_strength_duration_rationale is not None else "Insufficient data",
                'Probability Technical Success Rationale': result.probability_technical_success_rationale if result.probability_technical_success_rationale is not None else "Insufficient data",
                'Competitive Landscape Rationale': result.competitive_landscape_rationale if result.competitive_landscape_rationale is not None else "Insufficient data",
                'Transactability Rationale': result.transactability_rationale if result.transactability_rationale is not None else "Insufficient data",
                'Regulatory Path Complexity Rationale': result.regulatory_path_complexity_rationale if result.regulatory_path_complexity_rationale is not None else "Insufficient data",
                'Strategic Fit Rationale': result.strategic_fit_rationale if result.strategic_fit_rationale is not None else "Insufficient data"
            }
            
            if i == 0:  # Log first result for debugging
                self.logger.debug(f"First result scoring data: novelty={result.novelty_differentiation_score}, unmet_need={result.unmet_medical_need_score}")
                self.logger.debug(f"Output columns count: {len(self.output_columns)}")
                self.logger.debug(f"Row data keys count: {len(row.keys())}")
            
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _create_excel_from_csv(self, csv_file: Path, filtered_df: pd.DataFrame, output_path: Path) -> None:
        """Create Excel file from CSV with both original and filtered sheets."""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Add original data sheet
            original_df = pd.read_csv(csv_file)
            original_df.to_excel(writer, sheet_name='Original', index=False)
            
            # Add filtered data sheet
            filtered_df.to_excel(writer, sheet_name='Filtered', index=False)
            
            # Format the filtered sheet
            workbook = writer.book
            filtered_sheet = workbook['Filtered']
            self.format_worksheet(filtered_sheet, filtered_df)
    
    def _add_sheet_to_excel(self, excel_file: Path, filtered_df: pd.DataFrame) -> None:
        """Add filtered sheet to existing Excel file."""
        # Load existing workbook
        workbook = openpyxl.load_workbook(excel_file)
        
        # Remove existing 'Filtered' sheet if it exists
        if 'Filtered' in workbook.sheetnames:
            workbook.remove(workbook['Filtered'])
        
        # Create new filtered sheet
        filtered_sheet = workbook.create_sheet('Filtered')
        
        # Write headers
        for col_idx, column_name in enumerate(self.output_columns, 1):
            cell = filtered_sheet.cell(row=1, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.fill = self.colors['header']
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, (_, row) in enumerate(filtered_df.iterrows(), 2):
            for col_idx, column_name in enumerate(self.output_columns, 1):
                value = row[column_name]
                cell = filtered_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply formatting
        self.format_worksheet(filtered_sheet, filtered_df)
        
        # Save workbook
        workbook.save(excel_file)
        self.logger.info(f"Added 'Filtered' sheet to {excel_file}")
    
    def format_worksheet(self, worksheet, df: pd.DataFrame) -> None:
        """Apply conditional formatting to the worksheet."""
        try:
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Set minimum and maximum widths
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply conditional formatting to Pursue column (column E)
            pursue_col = 5  # Column E
            
            # Green for "Yes"
            yes_rule = CellIsRule(
                operator='equal',
                formula=['"Yes"'],
                fill=self.colors['pursue']
            )
            worksheet.conditional_formatting.add(
                f'E2:E{len(df) + 1}',
                yes_rule
            )
            
            # Red for "No"
            no_rule = CellIsRule(
                operator='equal',
                formula=['"No"'],
                fill=self.colors['dont_pursue']
            )
            worksheet.conditional_formatting.add(
                f'E2:E{len(df) + 1}',
                no_rule
            )
            
            # Yellow for "Error"
            error_rule = CellIsRule(
                operator='equal',
                formula=['"Error"'],
                fill=self.colors['error']
            )
            worksheet.conditional_formatting.add(
                f'E2:E{len(df) + 1}',
                error_rule
            )
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(self.output_columns)):
                for cell in row:
                    cell.border = thin_border
            
            # Freeze top row
            worksheet.freeze_panes = 'A2'
            
            self.logger.debug("Applied formatting to worksheet")
            
        except Exception as e:
            self.logger.warning(f"Failed to apply some formatting: {str(e)}")
    
    def generate_csv_output(self, results: List[ProcessingResult], output_path: Path) -> None:
        """Generate CSV output for dry-run mode."""
        try:
            df = self._results_to_dataframe(results)
            df.to_csv(output_path, index=False)
            self.logger.info(f"Generated CSV output: {output_path}")
        except Exception as e:
            raise SpreadsheetGeneratorError(f"Failed to generate CSV: {str(e)}")
    
    def get_summary_stats(self, results: List[ProcessingResult]) -> Dict[str, Any]:
        """Generate summary statistics for the results."""
        if not results:
            return {"total": 0, "message": "No results to analyze"}
        
        total_pairs = len(results)
        pursue_count = sum(1 for r in results if r.pursue is True)
        dont_pursue_count = sum(1 for r in results if r.pursue is False)
        error_count = sum(1 for r in results if r.pursue is None)
        
        repurposing_count = sum(1 for r in results if r.is_repurposing)
        primary_count = total_pairs - repurposing_count
        
        return {
            "total_asset_indication_pairs": total_pairs,
            "pursue_count": pursue_count,
            "dont_pursue_count": dont_pursue_count,
            "error_count": error_count,
            "pursue_rate_percent": round((pursue_count / total_pairs) * 100, 1),
            "error_rate_percent": round((error_count / total_pairs) * 100, 1),
            "primary_indications": primary_count,
            "repurposing_indications": repurposing_count
        }